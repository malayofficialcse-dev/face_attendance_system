"""
=========================================================
AI Face Attendance System
Main Module — v2.0

Features:
  - Check-In / Check-Out with hours tracking
  - Late Arrival Detection
  - Real Emotion Detection (DeepFace, background thread)
  - Multi-Camera Support (threaded)
  - Liveness Detection (MediaPipe eye-blink EAR)

Author : Malay Maity
=========================================================
"""

import cv2
import numpy as np
import pandas as pd
import pickle
import os
import time
import logging
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import math

import config
from email_report import send_attendance_report
from liveness import LivenessDetector
from emotion_detector import EmotionDetectorThread
from camera_thread import CameraThread

# =========================================================
# LOGGING SETUP
# =========================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(config.LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger("attendance")

# =========================================================
# CONFIGURATION ALIASES
# =========================================================

MODEL_FILE            = config.MODEL_FILE
LABELS_FILE           = config.LABELS_FILE
CSV_FILE              = config.CSV_FILE
EXCEL_FILE            = config.EXCEL_FILE
CONFIDENCE_THRESHOLD  = config.LBPH_CONFIDENCE_THRESHOLD
PROCESS_SCALE         = config.PROCESS_SCALE
FONT                  = cv2.FONT_HERSHEY_SIMPLEX
UNKNOWN_FOLDER        = config.UNKNOWN_FOLDER
SAVE_UNKNOWN_FACE     = config.SAVE_UNKNOWN_FACE
UNKNOWN_IMAGE_FORMAT  = config.UNKNOWN_IMAGE_FORMAT
UNKNOWN_PREFIX        = config.UNKNOWN_PREFIX
SHOW_CONFIDENCE       = config.SHOW_CONFIDENCE
FONT_SCALE            = config.FONT_SCALE
FONT_THICKNESS        = config.FONT_THICKNESS
BOX_THICKNESS         = config.BOX_THICKNESS
KNOWN_FACE_COLOR      = config.KNOWN_FACE_COLOR
UNKNOWN_FACE_COLOR    = config.UNKNOWN_FACE_COLOR
TEXT_COLOR            = config.TEXT_COLOR
INFO_COLOR            = config.INFO_COLOR
LATE_COLOR            = config.LATE_COLOR
WINDOW_NAME           = config.WINDOW_NAME
FRAME_WIDTH           = config.FRAME_WIDTH
FRAME_HEIGHT          = config.FRAME_HEIGHT
SKIP_FRAMES           = config.SKIP_FRAMES
SHOW_FPS              = config.SHOW_FPS
HAAR_CASCADE_PATH     = config.HAAR_CASCADE_PATH

PANEL_BG_COLOR        = (28, 28, 28)
PANEL_ACCENT_COLOR    = (12, 115, 255)
STATUS_BG_COLOR       = (10, 10, 10)

# =========================================================
# CSV / EXCEL COLUMNS (v2 schema)
# =========================================================

COLUMNS = ["Name", "Date", "Time", "Type", "Hours", "Late", "Emotion"]

# =========================================================
# CREATE / MIGRATE ATTENDANCE DIRECTORY & FILES
# =========================================================

if not os.path.exists(config.ATTENDANCE_FOLDER):
    os.makedirs(config.ATTENDANCE_FOLDER)


def _ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Add any missing v2 columns to a legacy DataFrame."""
    defaults = {
        "Name":    "",
        "Date":    "",
        "Time":    "",
        "Type":    "IN",
        "Hours":   "",
        "Late":    False,
        "Emotion": "Neutral",
    }
    for col, default in defaults.items():
        if col not in df.columns:
            df[col] = default
    return df[COLUMNS]


def create_attendance_files():
    # ---- CSV ----
    if not os.path.exists(CSV_FILE):
        pd.DataFrame(columns=COLUMNS).to_csv(CSV_FILE, index=False)
    else:
        try:
            df = pd.read_csv(CSV_FILE)
            df = _ensure_columns(df)
            df.to_csv(CSV_FILE, index=False)
        except Exception:
            pd.DataFrame(columns=COLUMNS).to_csv(CSV_FILE, index=False)

    # ---- Excel ----
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(COLUMNS)
        _style_header(ws)
        wb.save(EXCEL_FILE)
    else:
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            existing_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            for col in COLUMNS:
                if col not in existing_headers:
                    ws.cell(1, len(existing_headers) + 1, col)
                    existing_headers.append(col)
            wb.save(EXCEL_FILE)
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws.append(COLUMNS)
            _style_header(ws)
            wb.save(EXCEL_FILE)


def _style_header(ws):
    header_fill  = PatternFill("solid", fgColor="1C3A6E")
    header_font  = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center")
    for col_idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, col_idx)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = header_align


create_attendance_files()

# =========================================================
# LOAD LBPH MODEL
# =========================================================

log.info("=" * 60)
log.info("Loading Face Recognition Model...")

if not os.path.exists(MODEL_FILE) or not os.path.exists(LABELS_FILE):
    log.error("LBPH model or labels not found. Run encode_faces.py first.")
    exit(1)

recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read(MODEL_FILE)

with open(LABELS_FILE, "rb") as f:
    label_map = pickle.load(f)

reverse_labels = {v: k for k, v in label_map.items()}
log.info(f"Loaded {len(reverse_labels)} registered faces")

# =========================================================
# HELPER UTILITIES
# =========================================================

def current_date() -> str:
    return datetime.now().strftime(config.DATE_FORMAT)


def current_time_str() -> str:
    return datetime.now().strftime(config.TIME_FORMAT)


def face_confidence(raw_distance: float) -> float:
    return round(max(0.0, min(100.0, 100.0 - raw_distance)), 2)


def is_late() -> bool:
    """Return True if current time is past the configured late cutoff."""
    now = datetime.now()
    cutoff = now.replace(
        hour=config.LATE_ARRIVAL_HOUR,
        minute=config.LATE_ARRIVAL_MINUTE,
        second=0,
        microsecond=0,
    )
    return now > cutoff


# =========================================================
# ATTENDANCE DATA ACCESS
# =========================================================

_attendance_lock = threading.Lock()


def load_attendance_df() -> pd.DataFrame:
    if not os.path.exists(CSV_FILE):
        return pd.DataFrame(columns=COLUMNS)
    try:
        df = pd.read_csv(CSV_FILE)
        return _ensure_columns(df)
    except Exception:
        return pd.DataFrame(columns=COLUMNS)


def save_attendance_df(df: pd.DataFrame):
    df.to_csv(CSV_FILE, index=False)


def get_person_last_type(name: str, date: str) -> str | None:
    """
    Return the last attendance type ('IN' or 'OUT') for *name* on *date*,
    or None if no record exists.
    """
    df = load_attendance_df()
    records = df[(df["Name"] == name) & (df["Date"] == date)]
    if records.empty:
        return None
    return records.iloc[-1]["Type"]


def calculate_hours(name: str, date: str, checkout_time_str: str) -> str:
    """
    Find the most recent IN record for name/date and calculate elapsed hours.
    Returns a formatted string like "1h 23m" or "" if not found.
    """
    df = load_attendance_df()
    ins = df[(df["Name"] == name) & (df["Date"] == date) & (df["Type"] == "IN")]
    if ins.empty:
        return ""
    last_in_str = ins.iloc[-1]["Time"]
    try:
        fmt = config.TIME_FORMAT
        t_in  = datetime.strptime(last_in_str, fmt)
        t_out = datetime.strptime(checkout_time_str, fmt)
        delta = t_out - t_in
        total_seconds = int(delta.total_seconds())
        if total_seconds < 0:
            return ""
        h, rem = divmod(total_seconds, 3600)
        m, _   = divmod(rem, 60)
        return f"{h}h {m:02d}m"
    except Exception:
        return ""


# =========================================================
# MARK ATTENDANCE (Check-In / Check-Out aware)
# =========================================================

def mark_attendance(name: str, emotion: str = "Neutral") -> tuple[str, str]:
    """
    Mark attendance for *name*.

    Returns:
        (status_message, type_marked)  where type_marked is 'IN' or 'OUT'
    """
    date     = current_date()
    time_str = current_time_str()
    late     = is_late() if config.ENABLE_LATE_DETECTION else False

    with _attendance_lock:
        last_type = get_person_last_type(name, date)

        if config.ENABLE_CHECKIN_CHECKOUT:
            if last_type is None:
                mark_type = "IN"
            elif last_type == "IN":
                mark_type = "OUT"
            else:
                # Already OUT — start new IN cycle
                mark_type = "IN"
        else:
            # Legacy behaviour: no check-out; skip duplicates
            if last_type is not None and not config.ALLOW_DUPLICATE_ATTENDANCE:
                msg = f"{name} already marked today"
                log.info(msg)
                return msg, last_type

            mark_type = "IN"

        # Calculate hours for OUT marks
        hours = ""
        if mark_type == "OUT":
            hours = calculate_hours(name, date, time_str)

        # ---- write CSV ----
        df = load_attendance_df()
        new_row = pd.DataFrame([[name, date, time_str, mark_type, hours, late, emotion]],
                               columns=COLUMNS)
        df = pd.concat([df, new_row], ignore_index=True)
        save_attendance_df(df)

        # ---- write Excel ----
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

            # Ensure header exists
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            if not headers or headers[0] != "Name":
                ws.append(COLUMNS)

            row_data = [name, date, time_str, mark_type, hours, str(late), emotion]
            ws.append(row_data)

            # Color code OUT rows
            if mark_type == "OUT":
                row_num = ws.max_row
                fill = PatternFill("solid", fgColor="FFE0B2")  # light orange
                for col_idx in range(1, len(COLUMNS) + 1):
                    ws.cell(row_num, col_idx).fill = fill

            if late:
                row_num = ws.max_row
                late_fill = PatternFill("solid", fgColor="FFCCCC")  # light red
                for col_idx in range(1, len(COLUMNS) + 1):
                    ws.cell(ws.max_row, col_idx).fill = late_fill

            wb.save(EXCEL_FILE)
        except Exception as exc:
            log.warning(f"Excel write failed: {exc}")

    late_flag = " ⚠ LATE" if late else ""
    hours_str = f" | Hours: {hours}" if hours else ""
    msg = f"{mark_type}: {name} at {time_str}{late_flag}{hours_str}"
    log.info(msg)
    return msg, mark_type


# =========================================================
# ATTENDANCE COUNT
# =========================================================

def attendance_today_count() -> int:
    df = load_attendance_df()
    today = current_date()
    return len(df[(df["Date"] == today) & (df["Type"] == "IN")])


# =========================================================
# SAVE UNKNOWN FACE
# =========================================================

def save_unknown_face(frame, top, right, bottom, left):
    if not SAVE_UNKNOWN_FACE:
        return
    os.makedirs(UNKNOWN_FOLDER, exist_ok=True)
    timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
    face_image = frame[top:bottom, left:right]
    filepath = os.path.join(UNKNOWN_FOLDER, f"{UNKNOWN_PREFIX}{timestamp}{UNKNOWN_IMAGE_FORMAT}")
    cv2.imwrite(filepath, face_image)


# =========================================================
# UI HELPERS
# =========================================================

def draw_transparent_panel(frame, x, y, w, h, color, alpha=0.75):
    overlay = frame.copy()
    cv2.rectangle(overlay, (x, y), (x + w, y + h), color, cv2.FILLED)
    cv2.addWeighted(overlay, alpha, frame, 1 - alpha, 0, frame)


def draw_side_panel(frame, status_message, attendance_count, face_names,
                    face_types, liveness_status):
    panel_width  = 380
    panel_margin = 12
    x            = frame.shape[1] - panel_width - panel_margin
    y            = panel_margin
    panel_height = frame.shape[0] - panel_margin * 2

    draw_transparent_panel(frame, x, y, panel_width, panel_height, PANEL_BG_COLOR, 0.88)
    cv2.rectangle(frame, (x, y), (x + panel_width, y + 46), PANEL_ACCENT_COLOR, cv2.FILLED)
    cv2.putText(frame, "ATTENDANCE PANEL", (x + 16, y + 33),
                FONT, FONT_SCALE, TEXT_COLOR, FONT_THICKNESS, cv2.LINE_AA)

    cx        = x + 18
    line_y    = y + 74
    line_h    = int(30 * FONT_SCALE) + 6

    def put(text, color=TEXT_COLOR, scale_mult=1.0):
        nonlocal line_y
        cv2.putText(frame, text, (cx, line_y), FONT,
                    FONT_SCALE * scale_mult, color, 1, cv2.LINE_AA)
        line_y += line_h

    put(f"Status: {status_message[:28]}")
    put(f"Today (IN): {attendance_count}")

    line_y += 4
    put("Recent:", color=INFO_COLOR)
    recent = face_names[-4:] if face_names else ["None"]
    for i, name in enumerate(recent):
        t = face_types[-(len(recent)) + i] if face_types else ""
        put(f"  {name}  [{t}]", scale_mult=0.9)

    line_y += 6
    # Liveness
    lv_color = (0, 255, 0) if "VERIFIED" in liveness_status else (0, 80, 255)
    put(liveness_status[:35], color=lv_color, scale_mult=0.85)

    footer_y = y + panel_height - 24
    cv2.putText(frame, "Q=Quit  E=Email  S=Screenshot", (cx, footer_y),
                FONT, FONT_SCALE * 0.75, INFO_COLOR, 1, cv2.LINE_AA)


def draw_status_bar(frame, status_message):
    y = frame.shape[0] - 52
    draw_transparent_panel(frame, 0, y, frame.shape[1], 52, STATUS_BG_COLOR, 0.7)
    cv2.putText(frame, status_message, (14, y + 34),
                FONT, FONT_SCALE, TEXT_COLOR, FONT_THICKNESS, cv2.LINE_AA)
    cv2.putText(frame, "Center your face and avoid glare",
                (frame.shape[1] - 360, y + 34), FONT,
                FONT_SCALE * 0.85, INFO_COLOR, 1, cv2.LINE_AA)


def draw_face_box(frame, top, right, bottom, left,
                  name, confidence, color, emotion,
                  mark_type="IN", is_late_flag=False, liveness_verified=True):
    # Thin base rect
    cv2.rectangle(frame, (left, top), (right, bottom), color, 1)

    # Corner brackets
    length    = 20
    thickness = 4
    cv2.line(frame, (left,  top),    (left + length, top),    color, thickness)
    cv2.line(frame, (left,  top),    (left, top + length),    color, thickness)
    cv2.line(frame, (right, top),    (right - length, top),   color, thickness)
    cv2.line(frame, (right, top),    (right, top + length),   color, thickness)
    cv2.line(frame, (left,  bottom), (left + length, bottom), color, thickness)
    cv2.line(frame, (left,  bottom), (left, bottom - length), color, thickness)
    cv2.line(frame, (right, bottom), (right - length, bottom),color, thickness)
    cv2.line(frame, (right, bottom), (right, bottom - length),color, thickness)

    # Scanner animation
    box_height = bottom - top
    cycle      = (time.time() % 1.5) / 1.5
    progress   = cycle * 2 if cycle < 0.5 else (1.0 - cycle) * 2
    line_y     = int(top + progress * box_height)
    scan_color = (255, 255, 0)
    cv2.line(frame, (left, line_y),     (right, line_y),     scan_color, 2)
    cv2.line(frame, (left, line_y - 1), (right, line_y - 1), scan_color, 1)
    cv2.line(frame, (left, line_y + 1), (right, line_y + 1), scan_color, 1)

    # Emotion above box
    if emotion:
        cv2.putText(frame, f"Emotion: {emotion}",
                    (left, top - 26), FONT, 0.45, scan_color, 1, cv2.LINE_AA)

    # Late badge
    if is_late_flag and name != "Unknown":
        cv2.putText(frame, "LATE", (left, top - 10),
                    FONT, 0.55, LATE_COLOR, 2, cv2.LINE_AA)

    # Liveness warning
    if not liveness_verified:
        cv2.putText(frame, "BLINK TO VERIFY",
                    (left, bottom - 36), FONT, 0.45, (0, 80, 255), 2, cv2.LINE_AA)

    # Bottom label
    label = "Not Identified" if name == "Unknown" else f"{name} [{mark_type}]"
    if SHOW_CONFIDENCE and name != "Unknown":
        label += f" ({confidence:.0f}%)"

    label_bg_color = color
    cv2.rectangle(frame, (left, bottom), (right, bottom + 30), label_bg_color, cv2.FILLED)
    text_color = (0, 0, 0) if color == (0, 255, 0) else (255, 255, 255)
    cv2.putText(frame, label, (left + 6, bottom + 20),
                FONT, 0.5, text_color, 2, cv2.LINE_AA)


# =========================================================
# INITIALIZE SUBSYSTEMS
# =========================================================

log.info("Initializing subsystems...")

# Face cascade
face_cascade = cv2.CascadeClassifier(HAAR_CASCADE_PATH)

# Liveness detector (per-camera dict)
liveness_detectors: dict[int, LivenessDetector] = {}

# Emotion detector (shared background thread)
emotion_thread = EmotionDetectorThread()
if config.ENABLE_EMOTION_DETECTION:
    emotion_thread.start()
    log.info(f"Emotion detection: {'ON (DeepFace)' if emotion_thread.available else 'OFF (DeepFace unavailable)'}")
else:
    log.info("Emotion detection: disabled in config")

# =========================================================
# CAMERA THREADS
# =========================================================

log.info("Starting camera threads...")
camera_threads: dict[int, CameraThread] = {}

for cam_idx in config.CAMERA_INDICES:
    ct = CameraThread(cam_idx, FRAME_WIDTH, FRAME_HEIGHT)
    ct.start()
    camera_threads[cam_idx] = ct
    liveness_detectors[cam_idx] = LivenessDetector(
        ear_threshold=config.LIVENESS_EAR_THRESHOLD,
        consec_frames=config.LIVENESS_CONSEC_FRAMES,
        required_blinks=config.LIVENESS_REQUIRED_BLINKS,
    )

# Wait for cameras to open (retry loop inside CameraThread needs up to ~1s each)
time.sleep(3.0)

active_cameras = [idx for idx, ct in camera_threads.items() if ct.is_opened]
if not active_cameras:
    log.error("No cameras could be opened. Exiting.")
    exit(1)

log.info(f"Active cameras: {active_cameras}")
log.info("Press Q to quit | E to email report | S to screenshot")

# =========================================================
# PER-CAMERA STATE
# =========================================================

cam_state: dict[int, dict] = {}
for cam_idx in active_cameras:
    cam_state[cam_idx] = {
        "frame_counter":        0,
        "last_time":            time.time(),
        "prev_face_locations":  [],
        "prev_face_names":      [],
        "prev_face_confidences":[],
        "prev_face_emotions":   [],
        "prev_face_types":      [],
        "prev_late_flags":      [],
        "prev_status_message":  "No face detected",
        "liveness_result":      {"verified": not config.ENABLE_LIVENESS_DETECTION,
                                 "blinks": 0, "ear": 1.0,
                                 "status": "Liveness Disabled"},
        # track which names have been "unlocked" by liveness on this cam
        "liveness_unlocked":    set(),
        # track per-face liveness (reset when face disappears)
        "face_liveness":        {},   # name -> bool
        "emotion_frame_count":  0,
        "all_face_names":       [],   # running log for side panel
        "all_face_types":       [],
    }

# =========================================================
# MAIN LOOP
# =========================================================

running = True
while running:
    for cam_idx in active_cameras:
        ct    = camera_threads[cam_idx]
        state = cam_state[cam_idx]

        frame = ct.get_frame()
        if frame is None:
            continue

        h_frame, w_frame = frame.shape[:2]

        # ---------- Liveness (full-frame) ----------
        if config.ENABLE_LIVENESS_DETECTION:
            lv_result = liveness_detectors[cam_idx].process_frame(frame)
            state["liveness_result"] = lv_result
        else:
            state["liveness_result"] = {"verified": True, "blinks": 0,
                                        "ear": 1.0, "status": "Liveness Disabled"}

        # ---------- Face recognition (skip-frame) ----------
        should_process = (SKIP_FRAMES <= 1 or
                          state["frame_counter"] % SKIP_FRAMES == 0)

        face_locations  = state["prev_face_locations"]
        face_names      = state["prev_face_names"]
        face_confidences= state["prev_face_confidences"]
        face_emotions   = state["prev_face_emotions"]
        face_types      = state["prev_face_types"]
        late_flags      = state["prev_late_flags"]
        status_message  = state["prev_status_message"]

        if should_process:
            small_frame       = cv2.resize(frame, (0, 0),
                                           fx=PROCESS_SCALE, fy=PROCESS_SCALE)
            gray_small_frame  = cv2.cvtColor(small_frame, cv2.COLOR_BGR2GRAY)

            face_locations = face_cascade.detectMultiScale(
                gray_small_frame,
                scaleFactor=1.1,
                minNeighbors=5,
                minSize=(30, 30),
            )

            face_names       = []
            face_confidences = []
            face_emotions    = []
            face_types       = []
            late_flags       = []
            status_message   = "No face detected"

            for face_idx, (x, y, w, h) in enumerate(face_locations):
                # Scale crop for recognition
                face_gray = gray_small_frame[y:y + h, x:x + w]
                face_gray = cv2.resize(face_gray, (200, 200))
                face_gray = cv2.equalizeHist(face_gray)

                label, raw_conf = recognizer.predict(face_gray)
                confidence      = face_confidence(raw_conf)

                if raw_conf <= CONFIDENCE_THRESHOLD and label in reverse_labels:
                    name = reverse_labels[label]
                else:
                    name = "Unknown"

                # ---- Emotion (submit BGR crop to background thread) ----
                # Submit for ALL faces (known + unknown) so emotion always shows
                state["emotion_frame_count"] += 1
                if (config.ENABLE_EMOTION_DETECTION and
                        state["emotion_frame_count"] % config.EMOTION_SUBMIT_INTERVAL == 0):
                    ox_e = int(x / PROCESS_SCALE)
                    oy_e = int(y / PROCESS_SCALE)
                    ow_e = int(w / PROCESS_SCALE)
                    oh_e = int(h / PROCESS_SCALE)
                    face_bgr = frame[oy_e:oy_e + oh_e, ox_e:ox_e + ow_e]
                    if face_bgr.size > 0:
                        emotion_thread.submit(face_bgr, face_id=face_idx)

                emotion = emotion_thread.get_emotion(face_id=face_idx)

                # ---- Liveness gate ----
                lv_verified = state["liveness_result"]["verified"]

                if name != "Unknown":
                    already_unlocked = name in state["liveness_unlocked"]
                    if lv_verified and not already_unlocked:
                        state["liveness_unlocked"].add(name)
                    can_mark = already_unlocked or lv_verified
                else:
                    can_mark = True  # unknown always triggers save, not attendance

                # ---- Mark attendance ----
                mark_type = ""
                late_flag = False
                if name != "Unknown" and can_mark:
                    msg, mark_type = mark_attendance(name, emotion)
                    status_message = msg
                    late_flag = is_late() if config.ENABLE_LATE_DETECTION else False
                elif name == "Unknown":
                    status_message = "Unknown person detected"

                face_names.append(name)
                face_confidences.append(confidence)
                face_emotions.append(emotion)
                face_types.append(mark_type)
                late_flags.append(late_flag)

                if name != "Unknown":
                    state["all_face_names"].append(name)
                    state["all_face_types"].append(mark_type)

            # Save unknown faces
            for (x, y, w, h), name in zip(face_locations, face_names):
                if name == "Unknown":
                    ox = int(x / PROCESS_SCALE)
                    oy = int(y / PROCESS_SCALE)
                    ow = int(w / PROCESS_SCALE)
                    oh = int(h / PROCESS_SCALE)
                    save_unknown_face(frame, oy, ox + ow, oy + oh, ox)

            state["prev_face_locations"]   = face_locations
            state["prev_face_names"]       = face_names
            state["prev_face_confidences"] = face_confidences
            state["prev_face_emotions"]    = face_emotions
            state["prev_face_types"]       = face_types
            state["prev_late_flags"]       = late_flags
            state["prev_status_message"]   = status_message

        state["frame_counter"] += 1

        # =====================================================
        # DRAW RESULTS
        # =====================================================

        for i, (x, y, w, h) in enumerate(face_locations):
            ox = int(x / PROCESS_SCALE)
            oy = int(y / PROCESS_SCALE)
            ow = int(w / PROCESS_SCALE)
            oh = int(h / PROCESS_SCALE)

            name       = face_names[i]       if i < len(face_names)        else "Unknown"
            conf       = face_confidences[i] if i < len(face_confidences)  else 0
            emotion    = face_emotions[i]    if i < len(face_emotions)      else "Neutral"
            mark_type  = face_types[i]       if i < len(face_types)         else ""
            late_flag  = late_flags[i]       if i < len(late_flags)         else False
            lv_ok      = (name in state["liveness_unlocked"] or
                          state["liveness_result"]["verified"] or
                          not config.ENABLE_LIVENESS_DETECTION)

            if name == "Unknown":
                color = UNKNOWN_FACE_COLOR
            elif late_flag:
                color = LATE_COLOR
            else:
                color = KNOWN_FACE_COLOR

            draw_face_box(frame, oy, ox + ow, oy + oh, ox,
                          name, conf, color, emotion,
                          mark_type=mark_type,
                          is_late_flag=late_flag,
                          liveness_verified=lv_ok)

        # ---- Overlay labels ----
        cv2.putText(frame, f"Faces: {len(face_locations)}",
                    (10, 30), FONT, 0.8, (255, 255, 0), 2)
        cv2.putText(frame, f"Today IN: {attendance_today_count()}",
                    (10, 60), FONT, 0.7, (0, 255, 255), 2)

        # ---- Camera label (multi-cam) ----
        if len(active_cameras) > 1:
            cv2.putText(frame, f"CAM {cam_idx}", (10, 90),
                        FONT, 0.65, (200, 200, 200), 2, cv2.LINE_AA)

        draw_side_panel(frame, status_message,
                        attendance_today_count(),
                        state["all_face_names"][-6:],
                        state["all_face_types"][-6:],
                        state["liveness_result"]["status"])

        draw_status_bar(frame, status_message)

        liveness_detectors[cam_idx].draw_debug(frame, state["liveness_result"])

        # ---- FPS ----
        if SHOW_FPS:
            now_t = time.time()
            fps   = 1.0 / max(now_t - state["last_time"], 0.001)
            state["last_time"] = now_t
            cv2.putText(frame, f"FPS: {fps:.1f}",
                        (frame.shape[1] - 130, 30), FONT, 0.6, (0, 255, 0), 2)

        # ---- Date/time ----
        cv2.putText(frame, datetime.now().strftime(config.DATETIME_FORMAT),
                    (10, frame.shape[0] - 15), FONT, 0.6, (255, 255, 255), 2)

        # ---- Show window ----
        win_title = f"{WINDOW_NAME} — Camera {cam_idx}"
        cv2.imshow(win_title, frame)

    # =====================================================
    # KEY HANDLING (shared across all windows)
    # =====================================================

    key = cv2.waitKey(1) & 0xFF

    if key == ord("q"):
        log.info("Q pressed — shutting down...")
        running = False

    elif key == ord("e"):
        log.info("E pressed — sending email report...")
        send_attendance_report(CSV_FILE, EXCEL_FILE)

    elif key == ord("s"):
        ts = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        for cam_idx in active_cameras:
            ct    = camera_threads[cam_idx]
            frame = ct.get_frame()
            if frame is not None:
                path = os.path.join(config.ATTENDANCE_FOLDER,
                                    f"screenshot_cam{cam_idx}_{ts}.jpg")
                cv2.imwrite(path, frame)
                log.info(f"Screenshot saved: {path}")

# =========================================================
# CLEANUP
# =========================================================

log.info("Stopping camera threads...")
for ct in camera_threads.values():
    ct.stop()

emotion_thread.stop()

cv2.destroyAllWindows()

log.info("=" * 60)
log.info("Attendance System Closed Successfully")
log.info("=" * 60)